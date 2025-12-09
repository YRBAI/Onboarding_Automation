# excel_exporter.py - Enhanced Excel export with custom formatting

import pandas as pd
from config import EXCEL_COLUMNS, EXCEL_COLUMN_WIDTHS


class ExcelExporter:
    """Handles Excel export with enhanced formatting and text wrapping."""
    
    @staticmethod
    def export_to_excel(df: pd.DataFrame, filename: str) -> None:
        """Export DataFrame to Excel with enhanced formatting."""
        try:
            if df.empty:
                print("\n⚠️ No data to export.")
                return
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fund Data', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Fund Data']
                
                # Set column widths
                for col, width in EXCEL_COLUMN_WIDTHS.items():
                    worksheet.column_dimensions[col].width = width
                
                # Import styles
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                wrap_alignment = Alignment(wrap_text=True, vertical='top')
                
                # Apply header formatting
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    # Center align headers
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Add freeze panes at I1 (freeze first 8 columns: A through H)
                worksheet.freeze_panes = 'E1'
                
                # Set row height to 120 for all data rows (excluding header)
                for row_num in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_num].height = 120
                
                # Apply text wrapping for specific columns
                # Fund House (B), Fund Name (C), Asset (F), Geographic (G), Sector (H), 
                # Factsheet (K), Highlights PHS Link (L), Investment Objective (M), Key Risks for Investors (N), Other Risks (O)
                wrap_columns = ['B', 'C', 'F', 'G', 'H', 'K', 'L', 'M', 'N', 'O']
                
                # Center alignment for all cells
                center_alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')
                wrap_center_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for i, cell in enumerate(row):
                        # Convert 0-based index to column letter
                        col_letter = chr(65 + i)  # A=65, B=66, etc.
                        
                        if col_letter in wrap_columns:
                            cell.alignment = wrap_center_alignment  # Center + wrap text
                        else:
                            cell.alignment = center_alignment  # Just center
                                
                # Highlight blank cells in yellow
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        val = cell.value
                        if val is None or (isinstance(val, str) and val.strip() == ""):
                            cell.fill = yellow_fill
            
            print(f"\n✅ Data exported to: {filename}")
            print("  📋 Note: Blank cells are highlighted in yellow")
            print("  📐 Row height set to 120, text wrapping applied to specified columns")
            print("  🔒 Freeze panes set at D1 (first 3 columns frozen)")
            
        except ImportError:
            # Fallback to basic Excel export without formatting
            df.to_excel(filename, index=False)
            print(f"\n✅ Data exported to: {filename} (basic formatting)")
        except Exception as e:
            print(f"\n❌ Error exporting to Excel: {e}")
            # Fallback to CSV
            csv_filename = filename.replace('.xlsx', '.csv')
            df.to_csv(csv_filename, index=False)
            print(f"✅ Data exported to CSV instead: {csv_filename}")

    @staticmethod
    def generate_summary(df: pd.DataFrame) -> None:
        """Generate and print summary statistics with risk analysis."""
        if df.empty:
            return
        
        print(f"\n=== SUMMARY ===")
        print(f"Total Funds Processed: {len(df)}")
        
        # Risk extraction analysis
        if 'Key Risks for Investors' in df.columns and 'Other Risks' in df.columns:
            standard_risks_count = df['Key Risks for Investors'].apply(lambda x: bool(x and str(x).strip())).sum()
            other_risks_count = df['Other Risks'].apply(lambda x: bool(x and str(x).strip())).sum()
            
            print(f"\n📊 Risk Extraction Analysis:")
            print(f"  Funds with standard risks: {standard_risks_count}/{len(df)} ({standard_risks_count/len(df)*100:.1f}%)")
            print(f"  Funds with unclassified risks: {other_risks_count}/{len(df)} ({other_risks_count/len(df)*100:.1f}%)")
            
            # Analyze most common standard risks
            if standard_risks_count > 0:
                all_standard_risks = []
                for risks_str in df['Key Risks for Investors'].dropna():
                    if risks_str and str(risks_str).strip():
                        risks = [r.strip() for r in str(risks_str).split(';') if r.strip()]
                        all_standard_risks.extend(risks)
                
                if all_standard_risks:
                    from collections import Counter
                    risk_counts = Counter(all_standard_risks)
                    print(f"\n🎯 Most Common Standard Risks:")
                    for risk, count in risk_counts.most_common(10):
                        print(f"  {risk}: {count} funds ({count/len(df)*100:.1f}%)")
            
            # Show examples of unclassified risks
            if other_risks_count > 0:
                print(f"\n🔍 Examples of Unclassified Risks (Manual Review Needed):")
                examples_shown = 0
                for _, row in df.iterrows():
                    other_risks = row.get('Other Risks', '')
                    if other_risks and str(other_risks).strip() and examples_shown < 5:
                        isin = row.get('ISIN', 'Unknown')
                        risks = str(other_risks).split(';')[:2]  # Show first 2 risks
                        risks_preview = '; '.join([r.strip() for r in risks])
                        print(f"  {isin}: {risks_preview}...")
                        examples_shown += 1
                
                if other_risks_count > 5:
                    print(f"  ... and {other_risks_count - 5} more funds with unclassified risks")
        
        # Count blanks for each column
        print(f"\n📋 Blank Fields Count:")
        critical_fields = ['Asset', 'Retail/AI', 'Investment Objective', 'Key Risks for Investors', 'PSPL Risk Classification'] 
        for col in df.columns:
            blank_count = df[col].apply(lambda x: x == "" or pd.isna(x) or (isinstance(x, str) and not x.strip())).sum()
            if blank_count > 0 and col != 'No.':
                indicator = "⚠️" if col in critical_fields else "  "
                print(f"  {indicator} {col}: {blank_count} blanks")
        
        # Asset distribution
        asset_counts = df['Asset'].value_counts(dropna=False)
        if not asset_counts.empty:
            print(f"\n📈 Asset Distribution:")
            for asset, count in asset_counts.items():
                asset_display = asset if asset != "" and not pd.isna(asset) else "(Blank)"
                print(f"  {asset_display}: {count}")
        
        # Retail/AI distribution
        retail_ai_counts = df['Retail/AI'].value_counts(dropna=False)
        if not retail_ai_counts.empty:
            print(f"\n🏪 Retail/AI Distribution:")
            for classification, count in retail_ai_counts.items():
                classification_display = classification if classification != "" and not pd.isna(classification) else "(Blank - needs attention)"
                print(f"  {classification_display}: {count}")