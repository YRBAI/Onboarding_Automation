import pandas as pd
from rapidfuzz import fuzz
import openpyxl
from openpyxl.styles import PatternFill
import re


def jaccard_similarity(str1, str2):
    """
    Calculate Jaccard similarity between two strings based on word overlap.
    Returns a score between 0 and 100.
    """
    # Convert to lowercase and split into words
    words1 = set(str1.lower().split())
    words2 = set(str2.lower().split())
    
    # Calculate Jaccard similarity
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    
    if len(union) == 0:
        return 0.0
    
    return (len(intersection) / len(union)) * 100

def calculate_similarity_scores(fund_name, phillip_fund_name):
    """
    Calculate multiple similarity scores between two fund names.
    """
    # Handle missing values
    if pd.isna(fund_name) or pd.isna(phillip_fund_name):
        return {
            'Simple_Ratio': 0,
            'Partial_Ratio': 0,
            'Token_Sort_Ratio': 0,
            'Jaccard_Similarity': 0,
            'Average_Score': 0
        }
    
    # Convert to strings
    fund_name = str(fund_name)
    phillip_fund_name = str(phillip_fund_name)
    
    # Calculate fuzzy matching scores
    simple_ratio = fuzz.ratio(fund_name, phillip_fund_name)
    partial_ratio = fuzz.partial_ratio(fund_name, phillip_fund_name)
    token_sort_ratio = fuzz.token_sort_ratio(fund_name, phillip_fund_name)
    
    # Calculate Jaccard similarity
    jaccard_score = jaccard_similarity(fund_name, phillip_fund_name)
    
    # Calculate average score
    avg_score = (simple_ratio + partial_ratio + token_sort_ratio + jaccard_score) / 4
    
    return {
        'Simple_Ratio': round(simple_ratio, 2),
        'Partial_Ratio': round(partial_ratio, 2),
        'Token_Sort_Ratio': round(token_sort_ratio, 2),
        'Jaccard_Similarity': round(jaccard_score, 2),
        'Average_Score': round(avg_score, 2)
    }

def get_match_quality(avg_score):
    """
    Categorize match quality based on average score.
    """
    if avg_score >= 90:
        return "Excellent"
    elif avg_score >= 75:
        return "Good"
    elif avg_score >= 60:
        return "Fair"
    elif avg_score >= 40:
        return "Poor"
    else:
        return "Very Poor"

# Read the Excel file
input_file = 'FundName_SimilarityScore.xlsx'
df = pd.read_excel(input_file)

print(f"Processing {len(df)} fund records...")

# Calculate similarity scores for each row
similarity_scores = df.apply(
    lambda row: calculate_similarity_scores(row['Fund Name'], row['Phillip Fund Name']),
    axis=1
)

# Convert the similarity scores to a DataFrame
scores_df = pd.DataFrame(similarity_scores.tolist())

# Add the scores to the original DataFrame
df = pd.concat([df, scores_df], axis=1)

# Add Match Quality column
df['Match_Quality'] = df['Average_Score'].apply(get_match_quality)

# Reorder columns for better readability
columns_order = [
    'ISIN',
    'Fund Name',
    'Phillip Fund Name',
    'Simple_Ratio',
    'Partial_Ratio',
    'Token_Sort_Ratio',
    'Jaccard_Similarity',
    'Average_Score',
    'Match_Quality'
]
df = df[columns_order]

# Sort by Average Score (ascending) to see worst matches first
df_sorted = df.sort_values('Average_Score', ascending=True)

# Save to Excel with formatting
output_file = 'FundName_SimilarityScore_Results.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_sorted.to_excel(writer, sheet_name='Similarity Analysis', index=False)
    
    # Get the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Similarity Analysis']
    
    # Define color fills for match quality
    red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    green_fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
    
    # Apply conditional formatting based on Match Quality
    for idx, row in enumerate(df_sorted.itertuples(), start=2):  # start=2 to skip header
        match_quality = row.Match_Quality
        
        if match_quality == "Very Poor":
            worksheet[f'I{idx}'].fill = red_fill
        elif match_quality == "Poor":
            worksheet[f'I{idx}'].fill = yellow_fill
        elif match_quality == "Fair":
            worksheet[f'I{idx}'].fill = yellow_fill
        elif match_quality == "Good":
            worksheet[f'I{idx}'].fill = light_green_fill
        elif match_quality == "Excellent":
            worksheet[f'I{idx}'].fill = green_fill
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"\n✓ Analysis complete! Results saved to: {output_file}")

# Display summary statistics
print("\n" + "="*60)
print("SUMMARY STATISTICS")
print("="*60)
print(f"\nTotal records analyzed: {len(df)}")
print(f"\nMatch Quality Distribution:")
print(df['Match_Quality'].value_counts().sort_index())
print(f"\nAverage Similarity Scores:")
print(f"  Simple Ratio:       {df['Simple_Ratio'].mean():.2f}")
print(f"  Partial Ratio:      {df['Partial_Ratio'].mean():.2f}")
print(f"  Token Sort Ratio:   {df['Token_Sort_Ratio'].mean():.2f}")
print(f"  Jaccard Similarity: {df['Jaccard_Similarity'].mean():.2f}")
print(f"  Overall Average:    {df['Average_Score'].mean():.2f}")

# Show worst matches
print("\n" + "="*60)
print("TOP 10 WORST MATCHES (Lowest Similarity)")
print("="*60)
worst_matches = df_sorted.head(10)[['ISIN', 'Fund Name', 'Phillip Fund Name', 'Average_Score', 'Match_Quality']]
print(worst_matches.to_string(index=False))

# Show best matches
print("\n" + "="*60)
print("TOP 10 BEST MATCHES (Highest Similarity)")
print("="*60)
best_matches = df_sorted.tail(10)[['ISIN', 'Fund Name', 'Phillip Fund Name', 'Average_Score', 'Match_Quality']]
print(best_matches.to_string(index=False))

print("\n" + "="*60)
print(f"✓ Done! Check '{output_file}' for full results.")
print("="*60)